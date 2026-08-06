"""Lee una planilla de credenciales y saca una foto comparable.

La planilla del estudio no la diseñamos nosotros: ya existe, la mantiene gente
que no va a cambiar cómo trabaja, y va a tener celdas combinadas, filas en
blanco, columnas renombradas y una hoja llamada "Hoja1 (2)". Este módulo tiene
que tolerar todo eso sin romperse — un vigilante que falla cuando alguien
agrega una fila arriba es un vigilante que se apaga a la semana.

Decisión central: **las filas se identifican por su contenido, no por su
posición.** Si alguien ordena alfabéticamente o inserta una fila, todas las
posiciones cambian y un vigilante posicional gritaría que cambió todo. Acá la
identidad sale de las columnas que identifican la entrada (cliente + portal
por defecto), así que ordenar la planilla no genera ni un aviso.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path

# Columnas que, si aparecen, se tratan como secretas: su valor NUNCA se
# publica ni se guarda en el estado, sólo su hash. Se comparan en minúsculas y
# sin acentos para que "Clave", "clave fiscal" y "CLAVE" caigan todas acá.
PISTAS_SECRETO = ("clave", "contrasena", "contraseña", "password", "pass", "pin", "token")

# Columnas que identifican una fila. Si no están, se cae al hash de toda la fila.
PISTAS_IDENTIDAD = ("cliente", "empresa", "razon social", "razón social",
                    "portal", "sitio", "banco", "organismo", "usuario")


def _normalizar(texto: str) -> str:
    t = (texto or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        t = t.replace(a, b)
    return t


def es_secreta(columna: str) -> bool:
    n = _normalizar(columna)
    return any(p in n for p in PISTAS_SECRETO)


def es_identidad(columna: str) -> bool:
    n = _normalizar(columna)
    return any(p == n or p in n for p in PISTAS_IDENTIDAD)


def _huella(valor: str) -> str:
    """Hash del valor de un secreto.

    Se guarda el hash y no el valor para poder detectar que cambió sin tener
    nunca la clave en el archivo de estado. Si el estado se filtra, no entrega
    ni una credencial.
    """
    return hashlib.sha256((valor or "").encode()).hexdigest()[:16]


@dataclass(frozen=True, slots=True)
class Fila:
    hoja: str
    identidad: str                       # legible: "Pérez SA · ARCA"
    valores: dict[str, str]              # columna -> valor (secretos ya hasheados)
    secretas: frozenset[str] = frozenset()

    @property
    def clave(self) -> str:
        return f"{self.hoja}::{self.identidad}"


@dataclass(slots=True)
class Foto:
    """Estado de la planilla en un momento."""

    filas: dict[str, Fila] = field(default_factory=dict)
    hojas: tuple[str, ...] = ()

    def a_dict(self) -> dict:
        return {
            "hojas": list(self.hojas),
            "filas": {
                k: {"hoja": f.hoja, "identidad": f.identidad,
                    "valores": f.valores, "secretas": sorted(f.secretas)}
                for k, f in self.filas.items()
            },
        }

    @classmethod
    def desde_dict(cls, d: dict) -> Foto:
        filas = {
            k: Fila(hoja=v["hoja"], identidad=v["identidad"], valores=v["valores"],
                    secretas=frozenset(v.get("secretas", [])))
            for k, v in (d.get("filas") or {}).items()
        }
        return cls(filas=filas, hojas=tuple(d.get("hojas") or ()))


def _fila_vacia(valores: dict[str, str]) -> bool:
    return not any((v or "").strip() for v in valores.values())


def leer(ruta: Path | str, *, fila_encabezado: int = 1) -> Foto:
    """Lee un .xlsx y devuelve la foto. Los secretos salen hasheados.

    Recorre TODAS las hojas: en la planilla del estudio cada hoja es una
    sección (impuestos, bancos), que es justo el modelo de pestañas.
    """
    from openpyxl import load_workbook

    # read_only para no cargar en memoria una planilla grande, data_only para
    # obtener el resultado de las fórmulas y no el "=A1&B1".
    wb = load_workbook(filename=str(ruta), read_only=True, data_only=True)
    foto = Foto(hojas=tuple(wb.sheetnames))

    for hoja in wb.worksheets:
        filas = hoja.iter_rows(values_only=True)
        encabezado: list[str] = []
        for i, fila in enumerate(filas, start=1):
            if i < fila_encabezado:
                continue
            encabezado = [str(c).strip() if c is not None else "" for c in fila]
            break
        if not any(encabezado):
            continue  # hoja sin encabezado: se ignora en silencio

        secretas = frozenset(c for c in encabezado if c and es_secreta(c))
        cols_id = [c for c in encabezado if c and es_identidad(c)]

        for fila in filas:
            valores_crudos = {
                col: ("" if v is None else str(v).strip())
                for col, v in zip(encabezado, fila)
                if col
            }
            if _fila_vacia(valores_crudos):
                continue

            if cols_id:
                partes = [valores_crudos.get(c, "") for c in cols_id[:2]]
                identidad = " · ".join(p for p in partes if p)
            else:
                identidad = ""
            if not identidad:
                # Sin columnas de identidad utilizables, la fila se identifica
                # por el hash de sus valores no secretos. Editar un secreto
                # sigue detectándose; editar un campo visible se ve como
                # "se borró una y se agregó otra", que es aceptable y honesto.
                visible = {k: v for k, v in valores_crudos.items() if k not in secretas}
                identidad = "fila:" + _huella(repr(sorted(visible.items())))

            valores = {
                col: (_huella(v) if col in secretas else v)
                for col, v in valores_crudos.items()
            }
            f = Fila(hoja=hoja.title, identidad=identidad, valores=valores,
                     secretas=secretas)
            # Si dos filas tienen la misma identidad (duplicado en la planilla),
            # se desambigua para no perder ninguna.
            clave, n = f.clave, 2
            while clave in foto.filas:
                clave = f"{f.clave} ({n})"
                n += 1
            foto.filas[clave] = f

    wb.close()
    return foto
